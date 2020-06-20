import openpyxl
from openpyxl import load_workbook
from tqdm import tqdm
import sys

if __name__ == '__main__':
    excel_file = sys.argv[1]
    wb = load_workbook(excel_file, data_only = True)
    sheet_names = wb.get_sheet_names()

    free_space, obtacle, station_charging = 9, 1, 'FFFF0000'

    for sn in tqdm(sheet_names):
        sh = wb[sn]
        envioment = str(sh.max_row) + ' ' + str(sh.max_column) + '\n'
        for r in range(sh.max_row):
            for c in range(sh.max_column):
                color = sh.cell(r+1, c+1).fill.fgColor.value
                if color == free_space: envioment += '0 '
                elif color == obtacle : envioment += '1 '
                elif color == station_charging : envioment += '2 '
            envioment += '\n'
        with open(sn+'.txt', 'wt') as file:
            file.writelines(envioment)
